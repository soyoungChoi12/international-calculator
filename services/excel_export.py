"""국외여비지급내역서 Excel 생성.

원본 템플릿은 복사만 하고 수정하지 않는다.
openpyxl은 수식을 재계산하지 않으므로, 화면에 바로 보여야 하는 금액은 값으로 넣는다.
"""

from __future__ import annotations

import re
import shutil
import tempfile
from datetime import date
from io import BytesIO
from pathlib import Path

from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from config.excel_mapping import (
    BLANK_CELLS,
    EXCEL_CELL_MAP,
    FX_GUIDE_TEXT,
    FX_RATE_KIND,
    FX_SOURCE_TEXT,
    FX_SOURCE_URL,
    GRADE_CALC_COLUMNS,
    GRADE_CALC_ROWS,
    OUTPUT_SHEET_NAME,
    RATE_PLAIN_ROWS,
    TEMPLATE_PATH,
    TEMPLATE_SHEET_NAME,
    grade_calc_cell,
    lodging_check_cell,
)
from services.travel_calculator import RateSlice, TravelResult, lodging_actual_by_grade, truncate_to_ten


def excel_filename(traveler_name: str, when: date) -> str:
    safe = re.sub(r'[\\/:*?"<>|]', "", traveler_name.strip()) or "미기재"
    return f"국외여비지급내역서_{safe}_{when.strftime('%Y%m%d')}.xlsx"


def build_excel_bytes(result: TravelResult, traveler_name: str, approval_date: date) -> bytes:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Excel 템플릿을 찾을 수 없습니다: {TEMPLATE_PATH}")

    with tempfile.TemporaryDirectory() as tmp:
        work_path = Path(tmp) / TEMPLATE_PATH.name
        shutil.copy2(TEMPLATE_PATH, work_path)
        wb = load_workbook(work_path)
        _fill_workbook(wb, result, approval_date)
        buffer = BytesIO()
        wb.save(buffer)
        wb.close()
        return buffer.getvalue()


def _fill_workbook(wb, result: TravelResult, approval_date: date) -> None:
    if TEMPLATE_SHEET_NAME in wb.sheetnames:
        wb[TEMPLATE_SHEET_NAME].title = OUTPUT_SHEET_NAME
    ws = wb[OUTPUT_SHEET_NAME] if OUTPUT_SHEET_NAME in wb.sheetnames else wb[wb.sheetnames[0]]

    for address in BLANK_CELLS:
        ws[address] = None

    cells = EXCEL_CELL_MAP
    ws[cells["traveler_grade"]] = result.role_excel
    ws["C2"] = None
    ws[cells["fx_source_link"]] = FX_SOURCE_TEXT
    ws[cells["fx_source_link"]].hyperlink = FX_SOURCE_URL
    ws[cells["fx_guide"]] = FX_GUIDE_TEXT
    ws[cells["fx_rate"]] = result.exchange_rate
    ws[cells["fx_date_label"]] = f"{FX_RATE_KIND}({approval_date.strftime('%y.%m.%d')})"

    ws[cells["airfare_amount"]] = result.airfare_krw
    ws[cells["airfare_payment_method"]] = result.airfare_payment_method
    ws[cells["daily_amount"]] = result.daily.amount_krw
    ws[cells["daily_payment_method"]] = result.daily.payment_method
    ws[cells["meal_amount"]] = result.meal.amount_krw
    ws[cells["meal_payment_method"]] = result.meal.payment_method
    ws[cells["lodging_amount"]] = result.lodging.payable_krw
    ws[cells["lodging_payment_method"]] = result.lodging.payment_method
    ws[cells["preparation_amount"]] = result.preparation_krw
    ws[cells["preparation_payment_method"]] = result.preparation_payment_method
    ws[cells["total_amount"]] = result.total_krw
    ws[cells["corporate_card_total"]] = result.corporate_card_total
    ws[cells["personal_transfer_total"]] = result.personal_transfer_total
    ws[cells["payee_grand_total"]] = result.corporate_card_total + result.personal_transfer_total
    for address in (
        cells["airfare_amount"],
        cells["daily_amount"],
        cells["meal_amount"],
        cells["lodging_amount"],
        cells["preparation_amount"],
        cells["total_amount"],
        cells["corporate_card_total"],
        cells["personal_transfer_total"],
        cells["payee_grand_total"],
    ):
        ws[address].number_format = "#,##0"

    if result.rental_days:
        ws[cells["daily_note"]] = f"차량임차 {result.rental_days}일 일비 1/2 적용"
    if result.lodging.note:
        ws[cells["lodging_note"]] = result.lodging.note

    _fill_grade_blocks(ws, result)
    _clear_unused_grade_blocks(ws, result)
    _unstyled_rate_column(ws)
    _fill_lodging_check(ws, result)


def _fill_grade_blocks(ws: Worksheet, result: TravelResult) -> None:
    for grade, qty in result.grade_quantities.items():
        daily = _grade_line(result.daily.slices, grade, qty["daily"])
        lodging = _grade_line(result.lodging.slices, grade, qty["lodging"])
        meal = _grade_line(result.meal.slices, grade, qty["meal"])
        _write_grade_row(ws, grade, "daily", daily)
        _write_grade_row(ws, grade, "lodging", lodging)
        _write_grade_row(ws, grade, "meal", meal)


def _grade_line(
    slices: tuple[RateSlice, ...],
    grade: str,
    quantity: int,
) -> dict[str, int] | None:
    parts = [item for item in slices if item.grade == grade]
    if not parts and quantity <= 0:
        return None
    if not parts:
        return {"rate_usd": 0, "days": quantity, "amount_usd": 0, "amount_krw": 0}
    return {
        "rate_usd": max(item.rate_usd for item in parts),
        "days": quantity if quantity else sum(item.quantity for item in parts),
        "amount_usd": sum(item.amount_usd for item in parts),
        "amount_krw": sum(item.amount_krw for item in parts),
    }


def _write_grade_row(ws: Worksheet, grade: str, kind: str, line: dict[str, int] | None) -> None:
    if not line:
        return
    ws[grade_calc_cell(grade, kind, "rate_usd")] = line["rate_usd"]
    ws[grade_calc_cell(grade, kind, "days")] = line["days"]
    ws[grade_calc_cell(grade, kind, "amount_usd")] = line["amount_usd"]
    ws[grade_calc_cell(grade, kind, "amount_krw")] = line["amount_krw"]
    ws[grade_calc_cell(grade, kind, "amount_krw_rounded")] = truncate_to_ten(line["amount_krw"])
    ws[grade_calc_cell(grade, kind, "amount_krw")].number_format = "#,##0"
    ws[grade_calc_cell(grade, kind, "amount_krw_rounded")].number_format = "#,##0"


def _clear_unused_grade_blocks(ws: Worksheet, result: TravelResult) -> None:
    """출장에 없는 등급 칸의 조회 수식을 지운다. (예: 가급만 쓰면 I13에 식비 단가 49가 남는 것)"""
    used = set(result.grade_quantities)
    for grade, rows in GRADE_CALC_ROWS.items():
        if grade in used:
            continue
        for row in rows.values():
            for column in GRADE_CALC_COLUMNS.values():
                ws[f"{column}{row}"] = None


def _unstyled_rate_column(ws: Worksheet) -> None:
    """나·다·라 USD 단가 칸(I11:I19)을 가 등급과 같이 흰 배경·일반 글꼴로 맞춘다."""
    sample = ws["I7"].font
    white = PatternFill(fill_type="solid", fgColor="FFFFFF")
    for row in RATE_PLAIN_ROWS:
        cell = ws[f"I{row}"]
        cell.fill = white
        cell.font = Font(
            name=sample.name,
            size=sample.size,
            bold=False,
            italic=sample.italic,
            vertAlign=sample.vertAlign,
            underline=sample.underline,
            strike=sample.strike,
            color=copy(sample.color) if sample.color else None,
        )


def _fill_lodging_check(ws: Worksheet, result: TravelResult) -> None:
    nights_by_grade = [
        (grade, qty["lodging"])
        for grade, qty in result.grade_quantities.items()
        if qty["lodging"] > 0
    ]
    actual_by_grade = lodging_actual_by_grade(list(result.stays), result.lodging.actual_krw)
    ceiling_by_grade = {
        item.grade: item.amount_krw for item in result.lodging.slices if item.quantity > 0
    }
    for grade, nights in nights_by_grade:
        ceiling = ceiling_by_grade.get(grade, 0)
        actual = actual_by_grade.get(grade, 0)
        ws[lodging_check_cell(grade, "ceiling_krw")] = ceiling
        ws[lodging_check_cell(grade, "actual_krw")] = actual
        ws[lodging_check_cell(grade, "difference")] = ceiling - actual
        ws[lodging_check_cell(grade, "over_flag")] = "초과" if actual > ceiling else "미초과"
        ws[lodging_check_cell(grade, "final_krw")] = actual
        for key in ("ceiling_krw", "actual_krw", "difference", "final_krw"):
            ws[lodging_check_cell(grade, key)].number_format = "#,##0"
