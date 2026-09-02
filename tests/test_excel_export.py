"""STEP 5 Excel 산출 — 템플릿 사본에 계산 결과를 넣는다."""

from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from config.excel_mapping import OUTPUT_SHEET_NAME, TEMPLATE_PATH
from data.travel_rates import PAYMENT_CORPORATE, PAYMENT_PERSONAL
from services.excel_export import build_excel_bytes, excel_filename
from services.travel_calculator import StayInput, TravelInput, calculate_travel
from tests.test_travel_calculator import _sample


def _workbook(result, name: str = "홍길동", approval: date = date(2026, 8, 31)):
    data = build_excel_bytes(result, name, approval)
    return load_workbook(BytesIO(data)), data


def test_template_is_not_modified_by_export():
    before = TEMPLATE_PATH.read_bytes()
    result = calculate_travel(_sample())
    build_excel_bytes(result, "홍길동", date(2026, 8, 31))
    assert TEMPLATE_PATH.read_bytes() == before


def test_excel_filename_uses_name_and_date():
    assert excel_filename("홍길동", date(2026, 8, 31)) == "국외여비지급내역서_홍길동_20260831.xlsx"
    assert excel_filename("", date(2026, 9, 2)) == "국외여비지급내역서_미기재_20260902.xlsx"


def test_example_trip_fills_left_and_grade_a_block():
    result = calculate_travel(_sample())
    wb, _ = _workbook(result)
    assert OUTPUT_SHEET_NAME in wb.sheetnames
    ws = wb[OUTPUT_SHEET_NAME]
    assert ws["B2"].value == "그외직원(팀장및팀원)"
    assert ws["C2"].value is None
    assert ws["C6"].value == 1_200_000
    assert ws["D6"].value == PAYMENT_CORPORATE
    assert ws["C7"].value == 182_000
    assert ws["D7"].value == PAYMENT_PERSONAL
    assert ws["C8"].value == 469_000
    assert ws["C9"].value == 750_000
    assert ws["C10"].value == 70_000
    assert ws["C11"].value == 2_671_000
    assert ws["C18"].value == 2_020_000
    assert ws["C19"].value == 651_000
    assert ws["C20"].value == 2_671_000
    assert ws["L4"].value == 1400
    assert ws["M4"].value == "현찰 살 때(26.08.31)"
    assert ws["K3"].value == "하나은행 환율정보"
    assert ws["K3"].hyperlink.target == "https://www.kebhana.com/cont/mall/mall15/mall1501/index.jsp"
    assert "미국달러 현찰 살 때" in str(ws["M3"].value)
    assert ws["J7"].value == 5
    assert ws["I7"].value == 26
    assert ws["K7"].value == 130
    assert ws["L7"].value == 182_000
    assert ws["J8"].value == 4
    assert ws["I8"].value == 155
    assert ws["L8"].value == 868_000
    assert ws["J9"].value == 5
    assert ws["I9"].value == 67
    assert ws["L9"].value == 469_000
    assert ws["M7"].value == 182_000
    assert ws["M9"].value == 469_000
    assert ws["I13"].value is None
    assert ws["H26"].value == 868_000
    assert ws["I26"].value == 750_000
    assert ws["K26"].value == "미초과"
    assert ws["E6"].value is None
    assert ws["E9"].value is None
    assert ws["C12"].value is None
    wb.close()


def test_lodging_over_ceiling_writes_note():
    result = calculate_travel(_sample(lodging_actual_krw=1_000_000))
    wb, _ = _workbook(result)
    ws = wb[OUTPUT_SHEET_NAME]
    assert ws["C9"].value == 1_000_000
    assert "132,000원 초과" in str(ws["E9"].value)
    assert ws["K26"].value == "초과"
    wb.close()


def test_multi_city_fills_grade_a_and_b_day_cells():
    result = calculate_travel(
        _sample(
            lodging_nights=3,
            departure_date=date(2026, 8, 31),
            return_date=date(2026, 9, 3),
            lodging_actual_krw=500_000,
            stays=[
                StayInput("영국", "런던", 1, "가", stay_days=1),
                StayInput("영국", "버밍엄", 2, "나", stay_days=3),
            ],
        )
    )
    wb, _ = _workbook(result)
    ws = wb[OUTPUT_SHEET_NAME]
    assert ws["J7"].value == 1
    assert ws["J8"].value == 1
    assert ws["J9"].value == 1
    assert ws["J11"].value == 3
    assert ws["J12"].value == 2
    assert ws["J13"].value == 3
    assert ws["I13"].value == 49
    assert ws["C7"].value == 145_600
    wb.close()


def test_rental_days_write_daily_note_and_half_amount():
    result = calculate_travel(
        _sample(stays=[StayInput("미국", "샌프란시스코", 4, "가", stay_days=5, rental_days=2)])
    )
    wb, _ = _workbook(result)
    ws = wb[OUTPUT_SHEET_NAME]
    assert ws["C7"].value == 145_600
    assert ws["E7"].value == "차량임차 2일 일비 1/2 적용"
    assert ws["J7"].value == 5
    assert ws["K7"].value == 26 * 3 + 13 * 2
    wb.close()


def _is_white_fill(cell) -> bool:
    rgb = str(getattr(cell.fill.fgColor, "rgb", "") or "").upper()
    return rgb.endswith("FFFFFF")


def test_i11_i19_are_white_and_not_bold():
    result = calculate_travel(
        _sample(
            lodging_nights=3,
            departure_date=date(2026, 8, 31),
            return_date=date(2026, 9, 3),
            lodging_actual_krw=500_000,
            stays=[
                StayInput("영국", "런던", 1, "가", stay_days=1),
                StayInput("영국", "버밍엄", 2, "나", stay_days=3),
            ],
        )
    )
    wb, _ = _workbook(result)
    ws = wb[OUTPUT_SHEET_NAME]
    for row in range(11, 20):
        cell = ws[f"I{row}"]
        assert cell.font.bold is not True
        assert _is_white_fill(cell)
    wb.close()


def test_multi_city_lodging_actual_goes_to_grade_check():
    result = calculate_travel(
        _sample(
            lodging_nights=3,
            departure_date=date(2026, 8, 31),
            return_date=date(2026, 9, 3),
            lodging_actual_krw=0,
            stays=[
                StayInput("영국", "런던", 1, "가", stay_days=1, actual_krw=300_000),
                StayInput("영국", "버밍엄", 2, "나", stay_days=3, actual_krw=200_000),
            ],
        )
    )
    wb, _ = _workbook(result)
    ws = wb[OUTPUT_SHEET_NAME]
    assert ws["C9"].value == 500_000
    assert ws["I26"].value == 300_000
    assert ws["I27"].value == 200_000
    assert ws["L26"].value == 300_000
    assert ws["L27"].value == 200_000
    wb.close()


def test_fx_1523_writes_won_units_to_excel():
    result = calculate_travel(_sample(exchange_rate=1523))
    wb, _ = _workbook(result)
    ws = wb[OUTPUT_SHEET_NAME]
    assert ws["C7"].value == 197_990
    assert ws["C8"].value == 510_205
    assert ws["L7"].value == 197_990
    assert ws["L8"].value == 944_260
    assert ws["L9"].value == 510_205
    assert ws["M7"].value == 197_990
    assert ws["M9"].value == 510_200
    assert ws["H26"].value == 944_260
    assert ws["C19"].value == 708_195
    wb.close()
